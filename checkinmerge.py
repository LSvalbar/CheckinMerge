import os
import sys
import copy
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Alignment, PatternFill


class ExcelMerger(QWidget):
    def __init__(self):
        super().__init__()

        # 初始化UI
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel 文件合并工具")
        self.setGeometry(100, 100, 400, 200)

        self.layout = QVBoxLayout()

        # 标签显示合并结果
        self.label = QLabel("选择需要合并的Excel文件")
        self.layout.addWidget(self.label)

        # 选择文件按钮
        self.select_button = QPushButton("选择Excel文件")
        self.select_button.clicked.connect(self.select_files)
        self.layout.addWidget(self.select_button)

        # 合并文件按钮
        self.merge_button = QPushButton("开始合并")
        self.merge_button.clicked.connect(self.merge_excel_files)
        self.layout.addWidget(self.merge_button)

        self.setLayout(self.layout)

        # 存储选中的Excel文件路径
        self.selected_files = []

    def select_files(self):
        # 打开文件对话框选择多个Excel文件
        files, _ = QFileDialog.getOpenFileNames(self, "选择Excel文件", "", "Excel Files (*.xlsx);;All Files (*)")
        if files:
            self.selected_files = files
            self.label.setText(f"已选择 {len(files)} 个文件")

    def merge_excel_files(self):
        if not self.selected_files:
            self.label.setText("请先选择文件！")
            return

        # 创建一个新的Excel工作簿
        merged_wb = Workbook()
        merged_ws = merged_wb.active
        merged_ws.title = "考勤确认"

        first_file = self.selected_files[0]
        first_wb = load_workbook(first_file)
        first_ws = first_wb.active

        # 复制标题行及其格式
        for col_num, column in enumerate(first_ws.iter_cols(min_row=1, max_row=1), 1):
            merged_ws[get_column_letter(col_num) + '1'] = column[0].value
            merged_ws[get_column_letter(col_num) + '1'].font = copy.copy(column[0].font)
            merged_ws[get_column_letter(col_num) + '1'].border = copy.copy(column[0].border)
            merged_ws[get_column_letter(col_num) + '1'].fill = copy.copy(column[0].fill)
            merged_ws[get_column_letter(col_num) + '1'].alignment = copy.copy(column[0].alignment)

        # 复制列宽
        for col_num, col in enumerate(first_ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_num)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            merged_ws.column_dimensions[column].width = adjusted_width

        # 复制行高
        for row_num, row in enumerate(first_ws.iter_rows(), 1):
            merged_ws.row_dimensions[row_num].height = first_ws.row_dimensions[row_num].height

        current_row = 2
        # 遍历所有文件并合并数据
        for file in self.selected_files:
            count = 0
            wb = load_workbook(file)
            ws = wb.active

            # 从第二行开始复制内容及其格式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_num, cell in enumerate(row, 1):
                    new_cell = merged_ws[get_column_letter(col_num) + str(current_row)]
                    new_cell.value = cell.value
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.alignment = copy.copy(cell.alignment)
                    # if new_cell.column_letter == 'CD':
                    #     img = ws._images[count].ref
                    #     # 复制图片
                    #     # for image in ws._images:
                    #     #     img = Image(image.ref)
                    #     img.anchor = get_column_letter(col_num) + str(current_row)
                    #     img.width = 70
                    #     img.height = 23
                    #     # img.anchor = image.anchor
                    #     merged_ws.add_image(img)
                    #     count+=1
                current_row += 1

        # 保存合并后的文件
        output_file, _ = QFileDialog.getSaveFileName(self, "保存合并后的文件", "",
                                                     "Excel Files (*.xlsx);;All Files (*)")
        if output_file:
            merged_wb.save(output_file)
            self.label.setText(f"文件已保存为: {output_file}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelMerger()
    window.show()
    sys.exit(app.exec_())
